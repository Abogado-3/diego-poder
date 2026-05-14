#!/usr/bin/env python3
"""Búsqueda en el índice maestro de casos del firm.

La ruta del Excel del índice se lee de la config local (set por /diego-setup).

Subcomandos:
  buscar <query> [--top N]   → busca matches difusos
  por_radicado <rad>          → busca exacto por radicado
  decodificar <rad>           → solo decodifica el radicado (juzgado/ciudad)
  ruta                        → muestra la ruta resuelta (diagnóstico)
"""

import argparse
import json
import re
import sys
import unicodedata

import openpyxl

import diego_paths


# Códigos de ciudad (DANE/judicial - los más comunes)
CITIES = {
    "11001": "Bogotá D.C.",
    "76001": "Cali",
    "05001": "Medellín",
    "73001": "Ibagué",
    "08001": "Barranquilla",
    "66001": "Pereira",
    "63001": "Armenia",
    "50001": "Villavicencio",
    "68001": "Bucaramanga",
    "13001": "Cartagena",
    "17001": "Manizales",
    "19001": "Popayán",
    "23001": "Montería",
    "41001": "Neiva",
    "44001": "Riohacha",
    "47001": "Santa Marta",
    "52001": "Pasto",
    "54001": "Cúcuta",
    "70001": "Sincelejo",
    "85001": "Yopal",
    "20001": "Valledupar",
}

JURISDICCION = {
    "31": "del Circuito",
    "40": "Municipal",
    "33": "del Tribunal",
    "23": "Administrativo del Circuito",
    "41": "Administrativo del Circuito",
}

ESPECIALIDAD = {
    "03": "Civil",
    "04": "de Familia",
    "05": "Laboral",
    "06": "Penal",
    "07": "de Ejecución de Penas",
    "08": "Penal Especializado",
    "09": "de Ejecución de Garantías",
    "10": "Promiscuo",
    "11": "Promiscuo de Familia",
    "12": "de Restitución de Tierras",
    "18": "Pequeñas Causas",
    "39": "Pequeñas Causas y Competencia Múltiple",
    "89": "Pequeñas Causas y Competencia Múltiple",
}


def normalize(s) -> str:
    if not s:
        return ""
    s = str(s).lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def tokens(s: str) -> set:
    out = set()
    for t in normalize(s).split():
        if t.isdigit():
            if len(t) >= 3:
                out.add(t)
        elif len(t) >= 4:
            out.add(t)
    return out


def decodificar_radicado(radicado):
    if not radicado:
        return None
    rad = re.sub(r"\D", "", str(radicado))
    if len(rad) != 23:
        return {"valido": False, "raw": str(radicado), "digitos": len(rad)}

    ciudad_cod = rad[0:5]
    juris_cod = rad[5:7]
    esp_cod = rad[7:9]
    juzgado_num = int(rad[9:12])
    año = int(rad[12:16])
    consecutivo = int(rad[16:21])
    instancia = rad[21:23]

    ciudad = CITIES.get(ciudad_cod, f"[ciudad {ciudad_cod}]")
    juris = JURISDICCION.get(juris_cod, "")
    esp = ESPECIALIDAD.get(esp_cod, "")

    nombre_partes = [f"Juzgado {juzgado_num}", esp, juris, "de", ciudad]
    nombre = " ".join(p for p in nombre_partes if p).replace("  ", " ")

    return {
        "valido": True,
        "radicado_limpio": rad,
        "nombre_juzgado": nombre,
        "ciudad": ciudad,
        "ciudad_codigo": ciudad_cod,
        "jurisdiccion": juris,
        "jurisdiccion_codigo": juris_cod,
        "especialidad": esp,
        "especialidad_codigo": esp_cod,
        "numero_juzgado": juzgado_num,
        "año": año,
        "consecutivo": consecutivo,
        "instancia": instancia,
    }


def cargar_indice():
    index_path = diego_paths.indice_xlsx_path()
    if not index_path.exists():
        return []

    wb = openpyxl.load_workbook(index_path, data_only=True)
    cases = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        for r in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(r)]
            break
        if not headers:
            continue

        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not any(c for c in r if c):
                continue
            row_dict = {h: (str(v).strip() if v else "") for h, v in zip(headers, r)}
            row_dict["_hoja"] = sheet_name
            cases.append(row_dict)

    return cases


def extraer_radicado(case):
    for k, v in case.items():
        if not v or k.startswith("_"):
            continue
        kn = normalize(k)
        if "radicado" in kn:
            digits = re.sub(r"\D", "", str(v))
            if len(digits) >= 20:
                return digits, str(v)
    for k, v in case.items():
        if k.startswith("_"):
            continue
        digits = re.sub(r"\D", "", str(v or ""))
        if len(digits) == 23:
            return digits, str(v)
    return None, None


def enriquecer(case):
    rad_clean, rad_raw = extraer_radicado(case)
    enr = dict(case)
    enr["_radicado_limpio"] = rad_clean
    enr["_radicado_raw"] = rad_raw
    enr["_juzgado_deducido"] = decodificar_radicado(rad_clean) if rad_clean else None
    return enr


def cmd_buscar(args):
    q_tokens = tokens(args.query)
    cases = cargar_indice()

    q_digits = re.sub(r"\D", "", args.query)
    if len(q_digits) >= 18:
        for c in cases:
            rad, _ = extraer_radicado(c)
            if rad and (rad == q_digits or rad.endswith(q_digits) or q_digits.endswith(rad)):
                print(json.dumps({"matches": [enriquecer(c)], "via": "radicado_exacto"},
                                 ensure_ascii=False, indent=2))
                return

    scored = []
    for c in cases:
        haystack = " ".join(str(v) for k, v in c.items() if not k.startswith("_") and v)
        h_tokens = tokens(haystack)
        if not h_tokens or not q_tokens:
            continue
        hits = q_tokens & h_tokens
        if hits:
            score = len(hits) / max(len(q_tokens), 1)
            scored.append((score, c))

    scored.sort(key=lambda x: -x[0])
    top = [{"score": round(s, 3), **enriquecer(c)} for s, c in scored[: args.top]]
    print(json.dumps({"matches": top, "via": "tokens"}, ensure_ascii=False, indent=2))


def cmd_por_radicado(args):
    cases = cargar_indice()
    q = re.sub(r"\D", "", args.radicado)
    for c in cases:
        rad, _ = extraer_radicado(c)
        if rad and rad == q:
            print(json.dumps({"match": enriquecer(c)}, ensure_ascii=False, indent=2))
            return
    print(json.dumps({
        "match": None,
        "decodificado": decodificar_radicado(q),
    }, ensure_ascii=False, indent=2))


def cmd_decodificar(args):
    print(json.dumps(decodificar_radicado(args.radicado), ensure_ascii=False, indent=2))


def cmd_ruta(args):
    p = diego_paths.indice_xlsx_path()
    print(json.dumps({"path": str(p), "existe": p.exists()}, ensure_ascii=False))


def main():
    p = argparse.ArgumentParser(description="Búsqueda en índice maestro del firm")
    sub = p.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("buscar", help="Búsqueda difusa")
    b.add_argument("query")
    b.add_argument("--top", type=int, default=5)
    b.set_defaults(func=cmd_buscar)

    r = sub.add_parser("por_radicado", help="Match exacto por radicado")
    r.add_argument("radicado")
    r.set_defaults(func=cmd_por_radicado)

    d = sub.add_parser("decodificar", help="Solo decodifica un radicado")
    d.add_argument("radicado")
    d.set_defaults(func=cmd_decodificar)

    ru = sub.add_parser("ruta", help="Muestra dónde está el índice (diagnóstico)")
    ru.set_defaults(func=cmd_ruta)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
